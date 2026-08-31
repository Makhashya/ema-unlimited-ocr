r"""Single-file pipeline: nameplate images -> markdown -> CSV, over an API.

API-key twin of equipment_pipeline.py: the same three stages, but every model
call goes to a cloud (or local) vision-language model through an
OpenAI-compatible chat-completions API instead of the `claude` CLI.

  1. transcribe   the image is sent to the VLM and every character of text on
                  the nameplate is saved to  <stem>_text.md
  2. tabulate     the transcription is turned into one markdown equipment-
                  schedule table
  3. verify       the ORIGINAL image is sent again together with the draft
                  table; the model re-checks every cell, fixes OCR errors,
                  and fills blanks it can genuinely read or derive; the
                  verified table is saved to  <stem>_table.md
                  (skip this pass with --no-verify)

Finally all per-image tables are merged into one CSV. Outputs use the same
folders as equipment_pipeline.py, so the two are interchangeable:

    Mechanical_input\                                (your images)
    Output_Mechanical_input\
        output_text_Mechanical_input\   *_text.md
        output_table_Mechanical_input\  *_table.md
        output_csv_Mechanical_input\    Mechanical_input.csv

API setup is the same as vlm_to_md.py: keys live in environment variables or
in a `.env` file beside this script (template: .env.example). Real
environment variables win over .env values; command-line flags win over both.

    openai     https://api.openai.com/v1                            OPENAI_API_KEY
    anthropic  https://api.anthropic.com/v1  (OpenAI compat layer)  ANTHROPIC_API_KEY
    gemini     https://generativelanguage.googleapis.com/v1beta/openai  GEMINI_API_KEY
    custom     --base-url / VLM_BASE_URL (e.g. a local vLLM server)  API_KEY (optional)

Usage:
    python equipment_pipeline_api.py --image_dir Mechanical_input --provider anthropic
    python equipment_pipeline_api.py --image_dir Mechanical_input   (uses .env defaults)
    python equipment_pipeline_api.py --image_dir Mechanical_input --no-verify
    python equipment_pipeline_api.py --image_dir Mechanical_input --limit 2  # smoke test

Note: unlike the CLI version, the verify pass here has no live web search --
it re-reads the image and uses the model's own manufacturer knowledge only.
"""

import argparse
import base64
import csv
import os
import re
import sys
import time

import requests

EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}

# Formats every provider accepts as-is; anything else is re-encoded to JPEG.
RAW_OK = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
          ".webp": "image/webp"}

MAX_RAW_BYTES = 15 * 1024 * 1024   # stay under ~20 MB request limits

PROVIDERS = {
    # name: (base_url, key env var, default model)
    "openai":    ("https://api.openai.com/v1", "OPENAI_API_KEY", "gpt-4o"),
    "anthropic": ("https://api.anthropic.com/v1", "ANTHROPIC_API_KEY",
                  "claude-opus-5"),
    "gemini":    ("https://generativelanguage.googleapis.com/v1beta/openai",
                  "GEMINI_API_KEY", "gemini-2.5-flash"),
    "custom":    (None, "API_KEY", None),
}

MAX_RETRIES = 5
RETRYABLE = {429, 500, 502, 503, 504, 529}

COLUMNS = ("Tag #", "Make", "Model Number", "Serial Number", "Refrigerant Type",
           "Manufacture Date", "Approximate Tonnage", "MCA/MOCP",
           "Voltage/Phase", "Condition  (G/M/P)", "Comments")

TRANSCRIBE_PROMPT = """\
You are a meticulous transcription engine for equipment nameplate photos.
Read the attached image and transcribe ALL text visible in it, exactly as
printed.

Rules:
- Preserve every character of model numbers, serial numbers, part numbers,
  and electrical ratings exactly as shown. Never "correct", normalize, or
  guess a character. If a character is truly unreadable, write ? in its place.
- Keep each label with its value on one line, in the form "MODEL NO: <the
  characters printed on the plate>". Only ever write values you can actually
  see in the image.
- Transcribe printed, stamped, embossed, and handwritten text, including
  stickers and secondary labels.
- Output plain text only: no commentary, no descriptions of the image, no
  markdown code fences, no "Here is the transcription".
- If the image contains no readable text, output nothing."""

TABLE_PROMPT = """\
You extract HVAC equipment nameplate data from OCR text and return it as a \
single markdown table. You output the table and nothing else -- no preamble, \
no explanation, no closing remarks, no code fences.

The table has exactly these columns, in this order:

Tag # | Make | Model Number | Serial Number | Refrigerant Type | Manufacture \
Date | Approximate Tonnage | MCA/MOCP | Voltage/Phase | Condition  (G/M/P) | \
Comments

Rules:

- One row per physical unit found in the text. If the text describes a single \
nameplate, emit exactly one row.
- Emit a row whenever the text contains ANY equipment data at all -- a model \
number, serial number, manufacturer name, capacity, or electrical rating. \
Partial or messy OCR still gets a row, with the known cells filled and the \
rest blank. Supporting electrical equipment on the mechanical systems (a \
disconnect / safety switch, a unit heater, an exhaust fan) also gets a row: \
use the closest tag abbreviation and name the equipment type in Comments.
- Begin and end every row, including the header and separator, with a pipe.
- Tag #: use the equipment tag if the text gives one (e.g. CU5, AH2, RTU-3). \
If there is no tag, write the unit-type abbreviation followed by ( no tag ), \
e.g. "CU ( no tag )". Use these abbreviations: AH for an air handler, air \
handling unit, furnace, or indoor evaporator/cased coil; CU for a condensing \
unit or heat pump outdoor unit; RTU for a packaged rooftop unit; FC for a fan \
coil; B for a boiler; P for a pump; EF for an exhaust fan.
- Manufacture Date: the year. This is one of the MOST IMPORTANT columns -- \
the unit's age drives the whole assessment -- so never leave it blank \
without trying, in order: a date printed anywhere in the text, then \
decoding the serial number. If it is not stated outright but the serial \
number encodes it in the manufacturer's scheme, decode it -- e.g. Lennox \
serial 1519F31081 -> 2019 (week 15 of 2019), Ducane serial 1911G22987 -> \
2011. These are examples, not the only supported makes: apply the same idea \
to any manufacturer whose serial or tag convention you know (Trane, Carrier, \
York, Goodman, Rheem, smaller specialty makers, ...), including serials \
whose leading digits plainly read as a week-year, month-year, or year (e.g. \
02-02-... -> 2002). Fill the year whenever the decoding is clear; leave it \
blank only when the serial is genuinely ambiguous.
- Approximate Tonnage: whole tons, derived from the capacity digits in the \
model number when present -- e.g. -036- or 036 -> 3, 048 -> 4, 060 -> 5. \
Give the number only, with no unit.
- Refrigerant Type: copy the designation as printed, e.g. HFC-410A, HCFC-22.
- MCA/MOCP: combine as "MCA/MOCP" values separated by a slash, e.g. 18.6/30.
- Voltage/Phase: as printed, e.g. 208/230/1.
- Condition (G/M/P): only fill this in if the text actually assesses the \
unit's condition. Otherwise leave it blank.
- Leave any cell blank when the source text does not supply the value. Never \
invent, guess, or infer a value that is not supported by the text, and never \
write placeholders like N/A, unknown, or --.
- Preserve the OCR's exact characters for model and serial numbers. Do not \
"correct" them.

Example row (a Lennox air handler with no tag on it):

| AH ( no tag ) | Lennox | CX35-36A-6F-20 | 1519F31081 | HFC-410A | 2019 | 3 \
|  |  |  |  |

Only output the bare header and separator with no data rows when the text \
truly contains no equipment information of any kind (e.g. the transcription \
of a scenery photo)."""

VERIFY_PROMPT = TABLE_PROMPT + """

You are on a VERIFICATION pass. You are given the original nameplate photo \
(attached to the message), the raw transcription of it, and a draft table \
that was built from that transcription. Look at the photo yourself, then:

- Check every filled cell of the draft against what is actually printed in \
the image. Fix any cell the image contradicts -- especially model and serial \
numbers, where the draft may contain OCR character errors (0/O, 1/I, 5/S, \
8/B) or ? placeholders you can now resolve by looking closely at the plate.
- Hunt for missing information: for every blank cell, search the image \
(including stickers, secondary labels, and stamped text) for the value. Fill \
the cell only when you can genuinely read the value in the image or derive \
it under the rules above (year from the serial number, tonnage from the \
model number). A cell whose value simply is not present must stay blank.
- Give Manufacture Date special attention: if it is blank in the draft, \
look for a date printed or stamped anywhere in the image and decode the \
serial number's date scheme before accepting a blank cell.
- Do not drop or add rows unless the image clearly shows a different number \
of physical units than the draft has rows.
- Keep exactly the same columns, in the same order.

You may also use OUTSIDE INFORMATION -- your own knowledge of manufacturers \
-- to complete the verified plate data, under strict conditions:

- Use it to decode what the plate already gives you: the manufacturer's \
serial-number date scheme, the capacity/tonnage digits in the model number, \
the refrigerant a specific model line shipped with, or the full company name \
behind a logo or abbreviation on the plate.
- Only accept an outside fact when it is specific to THIS make and exact \
model/serial as read from the image, and the answer is unambiguous. If your \
knowledge is uncertain, covers a different model variant, or only a \
"similar" model, do not use it -- leave the cell as the plate supports.
- Never let remembered ratings (MCA, MOCP, voltage, refrigerant) override a \
cell the plate itself supplies; the plate always wins.
- Cells for values that are neither on the plate nor confidently derivable \
stay blank.

Output ONLY the corrected markdown table -- no commentary, no explanation of \
what you changed, no code fences."""

FENCE_RE = re.compile(r"^\s*```[^\n]*\n(.*?)\n\s*```\s*$", re.DOTALL)
SEP_CELL_RE = re.compile(r"^:?-{2,}:?$")
CELL_SPLIT_RE = re.compile(r"(?<!\\)\|")   # pipes that are not backslash-escaped


# --------------------------------------------------------------------------
# environment / API plumbing (same policy as vlm_to_md.py)
# --------------------------------------------------------------------------

def load_env_file():
    """Load KEY=VALUE lines from .env beside this script into os.environ.

    Values already set in the real environment are left alone, so exported
    variables always win over the file.
    """
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip().strip("'\"")
            if key and value:
                os.environ.setdefault(key, value)


def image_to_data_uri(path: str, max_side: int) -> str:
    """Base64 data URI for the image, re-encoding only when needed.

    Re-encode to JPEG when the format isn't universally accepted (bmp/tiff),
    the longest side exceeds max_side, or the file is too big for one request.
    """
    ext = os.path.splitext(path)[1].lower()
    mime = RAW_OK.get(ext)
    needs_reencode = mime is None or os.path.getsize(path) > MAX_RAW_BYTES

    if not needs_reencode and max_side > 0:
        from PIL import Image
        with Image.open(path) as im:
            needs_reencode = max(im.size) > max_side

    if needs_reencode:
        import io
        from PIL import Image
        with Image.open(path) as im:
            im = im.convert("RGB")
            if max_side > 0 and max(im.size) > max_side:
                scale = max_side / max(im.size)
                im = im.resize((round(im.width * scale), round(im.height * scale)),
                               Image.LANCZOS)
            buf = io.BytesIO()
            im.save(buf, format="JPEG", quality=92)
        data, mime = buf.getvalue(), "image/jpeg"
    else:
        with open(path, "rb") as f:
            data = f.read()

    return f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"


class ApiError(RuntimeError):
    pass


class PhotoTimeout(ApiError):
    """The per-photo time budget (--photo-timeout) ran out mid-stage."""


def strip_fence(text: str) -> str:
    m = FENCE_RE.match(text.strip())
    return m.group(1).strip() if m else text.strip()


def chat(session, base_url, headers, model, system_prompt, user_content,
         timeout, max_tokens, deadline=None) -> str:
    """One chat-completions call, with retry/backoff on transient errors.

    user_content is either a plain string or a list of content parts
    (text / image_url dicts) for vision requests. deadline, when given, is a
    time.perf_counter() value past which the call raises PhotoTimeout instead
    of starting (or retrying) a request.
    """
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ],
        "temperature": 0,
        "max_tokens": max_tokens,
    }
    last = None
    for attempt in range(MAX_RETRIES):
        eff_timeout = timeout
        if deadline is not None:
            budget = deadline - time.perf_counter()
            if budget <= 1:
                raise PhotoTimeout("photo time budget exceeded")
            eff_timeout = min(timeout, budget)
        try:
            resp = session.post(f"{base_url}/chat/completions", json=payload,
                                headers=headers, timeout=eff_timeout)
            if resp.status_code == 200:
                choice = resp.json()["choices"][0]
                content = (choice.get("message") or {}).get("content") or ""
                if content.strip():
                    return strip_fence(content)
                finish = choice.get("finish_reason")
                if finish != "length":
                    # an empty answer that stopped normally IS the answer --
                    # e.g. transcribing a photo with no readable text
                    return ""
                # A reasoning model can burn the whole token budget on its
                # hidden reasoning_content and return an EMPTY answer with
                # finish_reason "length". Retry with a larger budget so the
                # visible answer fits after the thinking.
                payload["max_tokens"] = min(payload["max_tokens"] * 4, 65536)
                last = (f"model returned empty content (finish_reason=length);"
                        f" retrying with max_tokens={payload['max_tokens']}")
                wait = 2
            else:
                body = resp.text[:300]
                if resp.status_code not in RETRYABLE:
                    # wrong model, bad key, oversized request:
                    # retrying won't help
                    raise ApiError(f"HTTP {resp.status_code}: {body}")
                last = f"HTTP {resp.status_code}: {body}"
                wait = 3 * (attempt + 1)
                if resp.status_code == 429 and resp.headers.get("Retry-After"):
                    try:
                        wait = min(float(resp.headers["Retry-After"]), 60.0)
                    except ValueError:
                        pass
        except requests.RequestException as e:
            last = f"{type(e).__name__}: {e}"
            wait = 3 * (attempt + 1)
        if attempt < MAX_RETRIES - 1:
            if deadline is not None:
                wait = min(wait, max(0.0, deadline - time.perf_counter()))
            time.sleep(wait)
    raise ApiError(f"gave up after {MAX_RETRIES} attempts: {last}")


# --------------------------------------------------------------------------
# stage 1: image -> transcription markdown
# --------------------------------------------------------------------------

def transcribe_image(api, image_path: str, max_side: int) -> str:
    uri = image_to_data_uri(image_path, max_side)
    content = [
        {"type": "image_url", "image_url": {"url": uri}},
        {"type": "text", "text": "Transcribe all text in this image, "
                                 "following your transcription rules."},
    ]
    return chat(api["session"], api["base_url"], api["headers"], api["model"],
                TRANSCRIBE_PROMPT, content, api["timeout"], api["max_tokens"],
                deadline=api.get("deadline"))


# --------------------------------------------------------------------------
# stage 2: transcription -> markdown table
# --------------------------------------------------------------------------

def add_image_column(table: str, image: str) -> str:
    """Prepend an 'Image' column as the first column of a markdown table.

    Done here rather than in the prompt so the value is exact on every row.
    """
    out, header_done = [], False
    for line in table.splitlines():
        s = line.strip()
        if not s.startswith("|"):
            out.append(line)
            continue
        cells = s.strip("|").split("|")
        non_empty = [c.strip() for c in cells if c.strip()]
        if not header_done:
            cells.insert(0, " Image ")
            header_done = True
        elif non_empty and all(SEP_CELL_RE.match(c) for c in non_empty):
            cells.insert(0, "---")
        else:
            cells.insert(0, f" {image} ")
        out.append("|" + "|".join(cells) + "|")
    return "\n".join(out)


def normalize_table(table: str) -> str:
    """Ensure the table starts with the canonical header and separator.

    Some models (especially smaller local VLMs) return only the data row(s);
    without a header the downstream parser would mistake the first data row
    for one. Detected by the absence of the 'Tag #' header cell.
    """
    lines = [ln for ln in table.splitlines() if ln.strip().startswith("|")]
    if not lines or any("Tag #" in ln for ln in lines):
        return table
    header = "| " + " | ".join(COLUMNS) + " |"
    sep = "|" + "---|" * len(COLUMNS)
    return "\n".join([header, sep, *lines])


def extract_table(api, transcription: str) -> str:
    table = normalize_table(
        chat(api["session"], api["base_url"], api["headers"], api["model"],
             TABLE_PROMPT, transcription, api["timeout"], api["max_tokens"],
             deadline=api.get("deadline")))
    if parse_table(table)[0] is None:
        # a header-only table is a legitimate "no data" answer, but NO table
        # at all means the model failed -- surface it instead of writing an
        # empty file that silently drops the transcription's information
        raise ApiError("the model returned no table for this transcription "
                       "(transcription is saved; retry or raise --max-tokens)")
    return table


# --------------------------------------------------------------------------
# stage 3: re-check the table against the original image
# --------------------------------------------------------------------------

def verify_table(api, image_path: str, transcription: str, draft_table: str,
                 max_side: int) -> str:
    """Second look: correct and complete the draft table from the image.

    Falls back to the draft when the verifier returns something that no
    longer parses as a table, so a bad verification pass can never lose data.
    """
    uri = image_to_data_uri(image_path, max_side)
    content = [
        {"type": "image_url", "image_url": {"url": uri}},
        {"type": "text", "text":
            f"Raw transcription of the attached photo:\n\n{transcription}\n\n"
            f"Draft table to verify and complete:\n\n{draft_table}"},
    ]
    checked = normalize_table(
        chat(api["session"], api["base_url"], api["headers"], api["model"],
             VERIFY_PROMPT, content, api["timeout"], api["max_tokens"],
             deadline=api.get("deadline")))
    header, rows = parse_table(checked)
    return checked if header and rows else draft_table


# --------------------------------------------------------------------------
# stage 4: markdown tables -> one CSV
# --------------------------------------------------------------------------

def split_row(line: str):
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [c.strip().replace("\\|", "|") for c in CELL_SPLIT_RE.split(s)]


def is_separator(cells) -> bool:
    non_empty = [c for c in cells if c]
    return bool(non_empty) and all(SEP_CELL_RE.match(c) for c in non_empty)


def parse_table(text: str):
    """Return (header, rows) from the first markdown table in the text."""
    header, rows = None, []
    for line in text.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = split_row(line)
        if header is None:
            header = cells
            continue
        if is_separator(cells):
            continue
        if any(c for c in cells):          # skip fully blank rows
            rows.append(cells)
    return header, rows


def tables_to_csv(table_dir: str, out_path: str, encoding: str = "utf-8-sig",
                  excel_path: str = None):
    files = sorted((f for f in os.listdir(table_dir)
                    if f.lower().endswith(".md")), key=natural_key)
    columns, records, empty = ["Image", *COLUMNS], [], []
    for name in files:
        with open(os.path.join(table_dir, name), encoding="utf-8") as f:
            header, rows = parse_table(f.read())
        if not header:
            print(f"  {name}: no table found, skipped")
            continue
        for col in header:                 # tolerate extra columns
            if col and col not in columns:
                columns.append(col)
        if not rows:
            empty.append(name)
            continue
        for cells in rows:
            # tolerate a row that is short or long relative to its header
            records.append({h: (cells[i] if i < len(cells) else "")
                            for i, h in enumerate(header) if h})

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        f = open(out_path, "w", newline="", encoding=encoding)
    except PermissionError:
        # the CSV is locked, usually because it is open in Excel -- don't
        # lose the run's work, write a timestamped copy beside it instead
        alt = out_path[:-len(".csv")] + time.strftime("_%H%M%S") + ".csv"
        print(f"cannot write {out_path} (locked -- open in Excel?); "
              f"writing {alt} instead")
        out_path = alt
        f = open(out_path, "w", newline="", encoding=encoding)
    with f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow(r)

    print(f"{len(records)} row(s) from {len(files) - len(empty)} table(s) "
          f"-> {out_path}")
    if empty:
        print(f"skipped in the CSV (no equipment data in the table): "
              f"{', '.join(empty)}")

    # -- Excel: one row per physical unit, all its photos listed together ----
    if excel_path and records:
        unique = dedupe_records(records, columns)
        try:
            written = write_excel(unique, columns, excel_path)
        except ImportError:
            print("openpyxl is not installed -- skipped the Excel file "
                  "(pip install openpyxl)")
        else:
            dups = len(records) - len(unique)
            print(f"{len(unique)} unit(s) -> {written}"
                  + (f"  ({dups} duplicate photo row(s) merged)" if dups
                     else ""))


# --------------------------------------------------------------------------
# stage 5: deduplicated Excel workbook
# --------------------------------------------------------------------------

def _cell(v) -> str:
    """A cell value as a clean string; '' for None/NaN (pandas gap-fill)."""
    if v is None or (isinstance(v, float) and v != v):
        return ""
    return str(v).strip()


def dedupe_records(records, columns):
    """Merge rows that describe the same physical unit.

    The same unit is often photographed more than once (wide shot + nameplate
    close-up). Rows sharing a non-blank Serial Number are one unit; rows
    without a serial merge only when every data cell except Image and
    Comments matches exactly. The merged row lists every source image in the
    Image cell, fills each blank cell from the duplicates, and unions the
    Comments.
    """
    data_cols = [c for c in columns if c not in ("Image", "Comments")]
    merged, index = [], {}
    for r in records:
        serial = _cell(r.get("Serial Number"))
        if serial:
            key = ("serial", serial.lower())
        else:
            key = ("row",) + tuple(_cell(r.get(c)).lower()
                                   for c in data_cols)
        hit = index.get(key)
        if hit is None:
            row = {c: _cell(r.get(c)) for c in columns}
            row["__images__"] = [_cell(r.get("Image"))]
            index[key] = row
            merged.append(row)
            continue
        image = _cell(r.get("Image"))
        if image and image not in hit["__images__"]:
            hit["__images__"].append(image)
        for c in columns:
            if c == "Image":
                continue
            v = _cell(r.get(c))
            if not v:
                continue
            cur = _cell(hit.get(c))
            if not cur:
                hit[c] = v
            elif c == "Comments" and v.lower() not in cur.lower():
                hit[c] = cur + "; " + v
    for row in merged:
        row["Image"] = ", ".join(i for i in row.pop("__images__") if i)
    return merged


def build_excel_workbook(records, columns):
    """One formatted 'Equipment List' sheet (openpyxl Workbook)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment List"
    ws.append(list(columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in records:
        ws.append([_cell(r.get(c)) for c in columns])
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    for i, col in enumerate(columns, 1):
        longest = max([len(str(col))]
                      + [len(_cell(r.get(col))) for r in records])
        ws.column_dimensions[get_column_letter(i)].width = \
            min(max(longest + 2, 10), 55)
    ws.freeze_panes = "A2"
    return wb


def write_excel(records, columns, out_path):
    wb = build_excel_workbook(records, columns)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        wb.save(out_path)
    except PermissionError:
        alt = out_path[:-len(".xlsx")] + time.strftime("_%H%M%S") + ".xlsx"
        print(f"cannot write {out_path} (locked -- open in Excel?); "
              f"writing {alt} instead")
        out_path = alt
        wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------
# glue
# --------------------------------------------------------------------------

def natural_key(s):
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", s)]


def paths_for(image_dir: str):
    """Every output path, from the input folder alone (project convention)."""
    src = os.path.abspath(image_dir)
    name = os.path.basename(src)
    root = os.path.join(os.path.dirname(src), f"Output_{name}")
    return {
        "name": name,
        "images": src,
        "root": root,
        "text": os.path.join(root, f"output_text_{name}"),
        "table": os.path.join(root, f"output_table_{name}"),
        "csv": os.path.join(root, f"output_csv_{name}", f"{name}.csv"),
        "excel": os.path.join(root, f"output_excel_{name}", f"{name}.xlsx"),
    }


def main():
    ap = argparse.ArgumentParser(
        description="nameplate images -> markdown -> equipment-schedule CSV, "
                    "over an OpenAI-compatible vision API (API-key twin of "
                    "equipment_pipeline.py)")
    ap.add_argument("--image_dir", required=True, help="folder of input images")
    ap.add_argument("--provider", choices=sorted(PROVIDERS), default=None,
                    help="which API preset to use "
                         "(default: VLM_PROVIDER from .env, else openai)")
    ap.add_argument("--base-url", default=None,
                    help="override the API base URL (required for --provider custom)")
    ap.add_argument("--model", default=None,
                    help="model name (default: the provider's preset)")
    ap.add_argument("--api-key-env", default=None,
                    help="name of the env var holding the API key "
                         "(default: the provider's usual variable)")
    ap.add_argument("--timeout", type=float, default=300.0,
                    help="per-attempt HTTP timeout in seconds (default: 300)")
    ap.add_argument("--photo-timeout", type=float, default=300.0,
                    help="total seconds allowed per photo across all stages; "
                         "on overrun the results extracted so far are saved "
                         "and the run moves to the next photo "
                         "(default: 300 = 5 min; 0 disables)")
    ap.add_argument("--max-tokens", type=int, default=16384,
                    help="response token cap per call (default: 16384; "
                         "reasoning models spend part of this thinking, and "
                         "an empty length-capped answer auto-retries larger)")
    ap.add_argument("--max-side", type=int, default=3000,
                    help="downscale images whose longest side exceeds this "
                         "many pixels (default: 3000; 0 disables)")
    ap.add_argument("--no-verify", action="store_true",
                    help="skip the verification pass that re-checks each "
                         "table against the original image (faster, cheaper)")
    ap.add_argument("--skip-existing", action="store_true",
                    help="reuse *_text.md / *_table.md files that already exist")
    ap.add_argument("--limit", type=int, default=0,
                    help="process only the first N images (0 = all)")
    args = ap.parse_args()

    if not os.path.isdir(args.image_dir):
        sys.exit(f"not a folder: {args.image_dir}")

    load_env_file()

    provider = args.provider or os.environ.get("VLM_PROVIDER", "").lower() or "openai"
    if provider not in PROVIDERS:
        ap.error(f"VLM_PROVIDER in .env must be one of "
                 f"{', '.join(sorted(PROVIDERS))} (got '{provider}')")

    # The VLM_* endpoint settings in .env describe one endpoint together with
    # VLM_PROVIDER, so they only apply when the provider wasn't chosen on the
    # command line -- an explicit --provider uses that provider's presets.
    def env_default(name):
        return None if args.provider else os.environ.get(name)

    preset_url, preset_env, preset_model = PROVIDERS[provider]
    base_url = (args.base_url or env_default("VLM_BASE_URL")
                or preset_url or "").rstrip("/")
    if not base_url:
        ap.error("--base-url (or VLM_BASE_URL in .env) is required "
                 "with the custom provider")
    model = args.model or env_default("VLM_MODEL") or preset_model
    if not model:
        ap.error("--model (or VLM_MODEL in .env) is required "
                 "with the custom provider")

    key_env = args.api_key_env or env_default("VLM_API_KEY_ENV") or preset_env
    api_key = os.environ.get(key_env, "")
    if not api_key and provider != "custom":
        sys.exit(f"no API key: set {key_env} in the environment or in "
                 f"{os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')} "
                 f"(template: .env.example)")

    api = {
        "session": requests.Session(),
        "base_url": base_url,
        "headers": {"Authorization": f"Bearer {api_key}"} if api_key else {},
        "model": model,
        "timeout": args.timeout,
        "max_tokens": args.max_tokens,
    }

    images = sorted((f for f in os.listdir(args.image_dir)
                     if os.path.splitext(f)[1].lower() in EXTS),
                    key=natural_key)
    if not images:
        sys.exit(f"no images found in {args.image_dir}")
    if args.limit > 0:
        images = images[:args.limit]

    p = paths_for(args.image_dir)
    os.makedirs(p["text"], exist_ok=True)
    os.makedirs(p["table"], exist_ok=True)
    print(f"input : {p['images']}  ({len(images)} image(s))")
    print(f"output: {p['root']}")
    print(f"api   : {provider}  model: {model}  ({base_url})")

    t0 = time.perf_counter()
    failed = []

    # Map each image to its output stems, disambiguating stems that collide
    # (ujju.png and ujju.jpg would otherwise both claim ujju_text.md).
    jobs, taken = [], {}
    for name in images:
        stem = os.path.splitext(name)[0]
        n = taken.get(stem.lower(), 0) + 1
        taken[stem.lower()] = n
        suffix = "" if n == 1 else f"_{n}"
        jobs.append((name,
                     os.path.join(p["text"], f"{stem}{suffix}_text.md"),
                     os.path.join(p["table"], f"{stem}{suffix}_table.md")))

    timed_out, skipped = [], []
    for i, (name, text_path, table_path) in enumerate(jobs, 1):
        t = time.perf_counter()
        image_path = os.path.join(p["images"], name)
        api["deadline"] = (t + args.photo_timeout) if args.photo_timeout > 0 \
            else None

        # -- stage 1: transcribe ---------------------------------------------
        try:
            if args.skip_existing and os.path.exists(text_path):
                text = open(text_path, encoding="utf-8").read().strip()
                note = "text reused"
            else:
                text = transcribe_image(api, image_path, args.max_side).strip()
                with open(text_path, "w", encoding="utf-8") as f:
                    f.write(text + "\n")
                note = f"{len(text)} chars"
        except PhotoTimeout:
            print(f"[{i}/{len(jobs)}] {name}  photo time budget exceeded "
                  f"during transcription -- nothing extracted, moving on")
            timed_out.append(name)
            continue
        except (ApiError, OSError) as e:
            print(f"[{i}/{len(jobs)}] {name}  FAILED: {e}")
            failed.append(name)
            continue
        if not text:
            print(f"[{i}/{len(jobs)}] {name}  no readable text -- image "
                  f"skipped")
            skipped.append(name)
            continue

        # -- stage 2: tabulate (+ stage 3: verify) ---------------------------
        try:
            if args.skip_existing and os.path.exists(table_path):
                note += ", table reused"
            else:
                table = extract_table(api, text)

                # -- stage 3: re-check against the original image ------------
                _, draft_rows = parse_table(table)
                if not args.no_verify and draft_rows:
                    try:
                        table = verify_table(api, image_path, text, table,
                                             args.max_side)
                        note += ", verified"
                    except PhotoTimeout:
                        # keep the unverified draft rather than lose it
                        note += ", verify timed out (draft table kept)"
                        timed_out.append(name)

                table = add_image_column(table, name)
                with open(table_path, "w", encoding="utf-8") as f:
                    f.write(f"# {name}\n\n{table}\n")
        except PhotoTimeout:
            print(f"[{i}/{len(jobs)}] {name}  photo time budget exceeded -- "
                  f"transcription saved, table skipped, moving on")
            timed_out.append(name)
            continue
        except (ApiError, OSError) as e:
            print(f"[{i}/{len(jobs)}] {name}  FAILED: {e}")
            failed.append(name)
            continue
        print(f"[{i}/{len(jobs)}] {name}  {time.perf_counter()-t:.1f}s  "
              f"{note} -> {os.path.basename(table_path)}")

    # -- stage 4+5: merge every table (including earlier runs) into the CSV
    # and the deduplicated Excel workbook --------------------------------
    print()
    tables_to_csv(p["table"], p["csv"], excel_path=p["excel"])

    print(f"\nDONE — {len(jobs) - len(failed)}/{len(jobs)} image(s) in "
          f"{time.perf_counter()-t0:.1f}s")
    if skipped:
        print(f"skipped (no readable text in the photo): "
              f"{', '.join(skipped)}")
    if timed_out:
        print(f"hit the {args.photo_timeout:.0f}s photo budget (partial "
              f"results saved; rerun with --skip-existing to finish): "
              f"{', '.join(timed_out)}")
    if failed:
        print(f"failed: {', '.join(failed)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
