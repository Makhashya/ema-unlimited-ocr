r"""Convert the merged equipment CSV into a formatted Excel workbook.

    output_csv_Mechanical_input\Mechanical_input.csv
        -> output_excel_Mechanical_input\Mechanical_input.xlsx

Usage:
    python csv_to_excel.py --csv output_csv_Mechanical_input\Mechanical_input.csv
    python csv_to_excel.py --csv_dir output_csv_Mechanical_input
    python csv_to_excel.py --csv ... --out D:\reports\schedule.xlsx

Requires openpyxl:  pip install openpyxl
"""

import argparse
import csv
import glob
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILL = PatternFill("solid", fgColor="F2F5FA")

WIDE_COLS = {"comments"}      # these get wrapped text and a fixed width
WIDE_WIDTH = 60
MIN_WIDTH, MAX_WIDTH = 10, 34


def xlsx_path_for(csv_path: str) -> str:
    """output_csv_X\\X.csv -> output_excel_X\\X.xlsx (beside the csv folder)."""
    p = os.path.abspath(csv_path)
    parent = os.path.dirname(p)
    stem = os.path.splitext(os.path.basename(p))[0]
    pname = os.path.basename(parent)
    if pname.startswith("output_csv_"):
        out_dir = os.path.join(os.path.dirname(parent),
                               "output_excel_" + pname[len("output_csv_"):])
    else:
        out_dir = os.path.join(parent, f"output_excel_{stem}")
    return os.path.join(out_dir, f"{stem}.xlsx")


def main():
    ap = argparse.ArgumentParser()
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--csv", help="path to the merged .csv")
    src.add_argument("--csv_dir", help="folder holding a single .csv, e.g. output_csv_ujju")
    ap.add_argument("--out", default=None,
                    help="output .xlsx (default: 'output_excel_<name>' beside the csv folder)")
    ap.add_argument("--sheet", default="Equipment", help="worksheet name")
    args = ap.parse_args()

    csv_path = args.csv
    if args.csv_dir:
        found = sorted(glob.glob(os.path.join(args.csv_dir, "*.csv")))
        if not found:
            sys.exit(f"no .csv found in {args.csv_dir}")
        if len(found) > 1:
            sys.exit(f"{len(found)} .csv files in {args.csv_dir} — pass --csv to pick one")
        csv_path = found[0]
    if not os.path.isfile(csv_path):
        sys.exit(f"not a file: {csv_path}")

    # utf-8-sig transparently handles both BOM and plain UTF-8
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    if not rows:
        sys.exit(f"{csv_path} is empty")

    header, data = rows[0], rows[1:]
    out_path = args.out or xlsx_path_for(csv_path)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet[:31]

    ws.append(header)
    for r in data:
        # pad short rows so borders/banding stay aligned
        ws.append(r + [""] * (len(header) - len(r)) if len(r) < len(header) else r)

    for c, name in enumerate(header, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    wide = {i for i, h in enumerate(header, 1) if h.strip().lower() in WIDE_COLS}
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=c in wide)
            if r % 2 == 0:
                cell.fill = BAND_FILL

    for c, name in enumerate(header, 1):
        letter = get_column_letter(c)
        if c in wide:
            ws.column_dimensions[letter].width = WIDE_WIDTH
            continue
        longest = max([len(name)] + [len(str(r[c - 1])) for r in data if c <= len(r)])
        ws.column_dimensions[letter].width = min(max(longest + 2, MIN_WIDTH), MAX_WIDTH)

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

    wb.save(out_path)
    print(f"{len(data)} row(s) x {len(header)} column(s) -> {out_path}")


if __name__ == "__main__":
    main()
